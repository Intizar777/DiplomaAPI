"""
Report generation service for production analytics with comprehensive logging.
"""
import csv
import io
from datetime import date, timedelta
from decimal import Decimal
from typing import List, Dict, Any, Optional, Literal

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc

from app.models import (
    AggregatedKPI,
    OrderSnapshot,
    QualityResult,
    AggregatedSales,
    SaleRecord,
    InventorySnapshot
)
from app.utils.logging_utils import track_feature_path, log_data_flow
import structlog

logger = structlog.get_logger()


class ReportService:
    """Service for generating production reports."""
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    @track_feature_path(feature_name="report.generate_report", log_result=True)
    async def generate_report(
        self,
        report_type: Literal["orders", "quality", "output", "sales", "inventory", "kpi"],
        format_type: Literal["csv", "xlsx"],
        date_from: date,
        date_to: date,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """Generate report in specified format."""
        filters = filters or {}
        
        logger.info(
            "generating_report",
            report_type=report_type,
            format_type=format_type,
            date_from=date_from,
            date_to=date_to
        )
        
        # Fetch data based on report type
        if report_type == "orders":
            data = await self._fetch_orders_data(date_from, date_to, filters)
        elif report_type == "quality":
            data = await self._fetch_quality_data(date_from, date_to, filters)
        elif report_type == "sales":
            data = await self._fetch_sales_data(date_from, date_to, filters)
        elif report_type == "inventory":
            data = await self._fetch_inventory_data(filters)
        elif report_type == "kpi":
            data = await self._fetch_kpi_data(date_from, date_to, filters)
        else:
            raise ValueError(f"Unknown report type: {report_type}")
        
        # Generate file
        if format_type == "csv":
            return self._generate_csv(data)
        elif format_type == "xlsx":
            return self._generate_xlsx(data, report_type)
        else:
            raise ValueError(f"Unknown format: {format_type}")
    
    async def _fetch_orders_data(
        self,
        date_from: date,
        date_to: date,
        filters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Fetch production orders data."""
        query = select(OrderSnapshot).where(
            OrderSnapshot.snapshot_date >= date_from,
            OrderSnapshot.snapshot_date <= date_to
        )
        
        if filters.get("status"):
            query = query.where(OrderSnapshot.status == filters["status"])
        if filters.get("production_line"):
            query = query.where(OrderSnapshot.production_line == filters["production_line"])
        
        query = query.order_by(desc(OrderSnapshot.snapshot_date))
        
        result = await self.db.execute(query)
        records = result.scalars().all()
        
        data = []
        for r in records:
            plan = r.target_quantity or Decimal("0")
            actual = r.actual_quantity or Decimal("0")
            completion = (actual / plan * 100) if plan > 0 else Decimal("0")
            
            data.append({
                "order_id": str(r.order_id),
                "external_id": r.external_order_id,
                "product": r.product_name or str(r.product_id),
                "line": r.production_line,
                "status": r.status,
                "planned_qty": float(plan),
                "actual_qty": float(actual),
                "completion_pct": float(completion.quantize(Decimal("0.1"))),
                "planned_start": r.planned_start.isoformat() if r.planned_start else "",
                "planned_end": r.planned_end.isoformat() if r.planned_end else "",
                "snapshot_date": r.snapshot_date.isoformat()
            })

        log_data_flow(
            source="database",
            target="report",
            operation="fetch",
            payload_summary={"order_fields": len(data[0].keys()) if data else 0},
            records_count=len(data),
        )
        return data

    async def _fetch_quality_data(
        self,
        date_from: date,
        date_to: date,
        filters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Fetch quality control data."""
        query = select(QualityResult).where(
            QualityResult.test_date >= date_from,
            QualityResult.test_date <= date_to
        )
        
        if filters.get("decision"):
            query = query.where(QualityResult.decision == filters["decision"])
        
        query = query.order_by(desc(QualityResult.test_date))
        
        result = await self.db.execute(query)
        records = result.scalars().all()
        
        data = []
        for r in records:
            data.append({
                "lot_number": r.lot_number,
                "product": r.product_name or str(r.product_id),
                "parameter": r.parameter_name,
                "value": float(r.result_value) if r.result_value else None,
                "lower_limit": float(r.lower_limit) if r.lower_limit else None,
                "upper_limit": float(r.upper_limit) if r.upper_limit else None,
                "in_spec": "Да" if r.in_spec else "Нет",
                "decision": r.decision,
                "test_date": r.test_date.isoformat()
            })

        log_data_flow(
            source="database",
            target="report",
            operation="fetch",
            payload_summary={"quality_fields": len(data[0].keys()) if data else 0},
            records_count=len(data),
        )
        return data

    async def _fetch_sales_data(
        self,
        date_from: date,
        date_to: date,
        filters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Fetch sales data."""
        query = select(SaleRecord).where(
            SaleRecord.sale_date >= date_from,
            SaleRecord.sale_date <= date_to
        )
        
        if filters.get("region"):
            query = query.where(SaleRecord.region == filters["region"])
        
        query = query.order_by(desc(SaleRecord.sale_date))
        
        result = await self.db.execute(query)
        records = result.scalars().all()
        
        data = []
        for r in records:
            data.append({
                "external_id": r.external_id,
                "product": r.product_name or str(r.product_id),
                "customer": r.customer_name,
                "quantity": float(r.quantity) if r.quantity else 0,
                "amount": float(r.amount) if r.amount else 0,
                "region": r.region,
                "channel": r.channel,
                "sale_date": r.sale_date.isoformat()
            })

        log_data_flow(
            source="database",
            target="report",
            operation="fetch",
            payload_summary={"sales_fields": len(data[0].keys()) if data else 0},
            records_count=len(data),
        )
        return data

    async def _fetch_inventory_data(
        self,
        filters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Fetch current inventory snapshot."""
        # Get latest snapshot date
        latest_date_query = select(func.max(InventorySnapshot.snapshot_date))
        latest_date_result = await self.db.execute(latest_date_query)
        latest_date = latest_date_result.scalar()
        
        if not latest_date:
            return []
        
        query = select(InventorySnapshot).where(
            InventorySnapshot.snapshot_date == latest_date
        )
        
        if filters.get("warehouse_code"):
            query = query.where(InventorySnapshot.warehouse_code == filters["warehouse_code"])
        
        result = await self.db.execute(query)
        records = result.scalars().all()
        
        data = []
        for r in records:
            days_to_expiry = None
            if r.expiry_date:
                days_to_expiry = (r.expiry_date - date.today()).days
            
            data.append({
                "product": r.product_name or str(r.product_id),
                "warehouse": r.warehouse_code,
                "quantity": float(r.quantity) if r.quantity else 0,
                "unit": r.unit_of_measure,
                "production_date": r.production_date.isoformat() if r.production_date else "",
                "expiry_date": r.expiry_date.isoformat() if r.expiry_date else "",
                "days_to_expiry": days_to_expiry,
                "status": "Критический" if r.quantity and r.quantity < 10 else "Норма"
            })

        log_data_flow(
            source="database",
            target="report",
            operation="fetch",
            payload_summary={"inventory_fields": len(data[0].keys()) if data else 0},
            records_count=len(data),
        )
        return data

    async def _fetch_kpi_data(
        self,
        date_from: date,
        date_to: date,
        filters: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """Fetch KPI data."""
        query = select(AggregatedKPI).where(
            AggregatedKPI.period_from >= date_from,
            AggregatedKPI.period_to <= date_to
        )
        
        if filters.get("production_line"):
            query = query.where(AggregatedKPI.production_line == filters["production_line"])
        
        query = query.order_by(desc(AggregatedKPI.period_from))
        
        result = await self.db.execute(query)
        records = result.scalars().all()
        
        data = []
        for r in records:
            data.append({
                "period_from": r.period_from.isoformat(),
                "period_to": r.period_to.isoformat(),
                "line": r.production_line or "Все линии",
                "total_output": float(r.total_output) if r.total_output else 0,
                "defect_rate_pct": float(r.defect_rate) if r.defect_rate else 0,
                "completed_orders": r.completed_orders,
                "total_orders": r.total_orders,
                "oee_pct": float(r.oee_estimate) if r.oee_estimate else None
            })

        log_data_flow(
            source="database",
            target="report",
            operation="fetch",
            payload_summary={"kpi_fields": len(data[0].keys()) if data else 0},
            records_count=len(data),
        )
        return data

    def _generate_csv(self, data: List[Dict[str, Any]]) -> bytes:
        """Generate CSV from data."""
        if not data:
            return b""
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        
        return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    
    def _generate_xlsx(self, data: List[Dict[str, Any]], report_type: str) -> bytes:
        """Generate Excel file from data."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed, falling back to CSV")
            return self._generate_csv(data)
        
        if not data:
            data = [{"message": "Нет данных для отчёта"}]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт"
        
        # Headers
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Style header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        for row_data in data:
            row = [row_data.get(h, "") for h in headers]
            ws.append(row)
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add title row
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = f"Отчёт: {report_type}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    async def get_report_metadata(self) -> List[Dict[str, Any]]:
        """Get available report types with metadata."""
        return [
            {
                "id": "orders",
                "name": "Производственные заказы",
                "description": "Статус и выполнение производственных заданий",
                "formats": ["csv", "xlsx"],
                "default_period": "week"
            },
            {
                "id": "quality",
                "name": "Контроль качества",
                "description": "Результаты тестирования и дефекты",
                "formats": ["csv", "xlsx"],
                "default_period": "week"
            },
            {
                "id": "sales",
                "name": "Продажи",
                "description": "Данные по реализации продукции",
                "formats": ["csv", "xlsx"],
                "default_period": "month"
            },
            {
                "id": "inventory",
                "name": "Остатки на складах",
                "description": "Текущие запасы ГП и сроки годности",
                "formats": ["csv", "xlsx"],
                "default_period": "day"
            },
            {
                "id": "kpi",
                "name": "KPI производства",
                "description": "Сводные показатели эффективности",
                "formats": ["csv", "xlsx"],
                "default_period": "month"
            }
        ]
